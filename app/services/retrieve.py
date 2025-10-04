# app/services/retrieve.py
from __future__ import annotations
import os
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from functools import lru_cache
from datetime import datetime

EXCEL_PATH = os.getenv("FAQ_EXCEL_PATH", "./data/เทสบอท รอบท้าย.xlsx")

# -----------------------------
# Utilities
# -----------------------------
def _sheet_to_dicts(ws) -> List[Dict[str, Optional[str]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    items: List[Dict[str, Optional[str]]] = []
    for r in rows[1:]:
        d: Dict[str, Optional[str]] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        items.append(d)
    return items

def _find_first_key(d: Dict[str, object], candidates: List[str]) -> Optional[str]:
    """หา key แรกที่มีอยู่ใน dict จากรายชื่อไทยหลายแบบ"""
    low = {k.lower(): k for k in d.keys()}
    for name in candidates:
        if name in d:
            return name
        if name.lower() in low:
            return low[name.lower()]
    return None

def safe_str(x: object) -> str:
    return "" if x is None else str(x).strip()

# -----------------------------
# Load workbook with cache
# -----------------------------
@lru_cache(maxsize=1)
def _load_workbook_cached(mtime: float) -> Tuple[object, float]:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    return wb, mtime

def _workbook():
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
        wb, _ = _load_workbook_cached(mtime)
        return wb
    except Exception:
        # ถ้าไฟล์ยังไม่พร้อม ให้โยนให้ caller จัดการข้อความสุภาพ
        raise

# -----------------------------
# Public APIs ที่ respond.py เรียกใช้
# -----------------------------
def get_faq_answer(query: str) -> Optional[str]:
    """
    อ่านชีท 'FAQ' แล้วหาคำตอบจากคอลัมน์คีย์เวิร์ด/คำถาม/คำตอบ
    ชื่อหัวคอลัมน์ยืดหยุ่น: ['คีย์เวิร์ด', 'คีย์เวิรด์', 'keyword'], ['คำถาม'], ['คำตอบ']
    """
    wb = _workbook()
    if "FAQ" not in wb.sheetnames:
        return None
    ws = wb["FAQ"]
    items = _sheet_to_dicts(ws)
    ans_col = _find_first_key(items[0] if items else {}, ["คำตอบ", "answer"])
    kw_col = _find_first_key(items[0] if items else {}, ["คีย์เวิร์ด", "คีย์เวิรด์", "keyword"])
    q_col  = _find_first_key(items[0] if items else {}, ["คำถาม", "question"])
    if not ans_col:
        return None

    q = query.lower()
    # 1) match คีย์เวิร์ดก่อน
    if kw_col:
        for row in items:
            kws = safe_str(row.get(kw_col))
            if not kws:
                continue
            for kw in [k.strip() for k in kws.split(",")]:
                if kw and kw.lower() in q:
                    return safe_str(row.get(ans_col))
    # 2) ย้อนมาดูคำถามแบบ contains
    if q_col:
        for row in items:
            question = safe_str(row.get(q_col)).lower()
            if question and any(tok in q for tok in question.split()):
                return safe_str(row.get(ans_col))
    return None

def get_product_info(query: str) -> Optional[str]:
    """
    อ่านชีท 'ข้อมูลสินค้าและราคา'
    หาจากชื่อสินค้า / alias แล้วประกอบข้อความตอบ
    คอลัมน์ที่รองรับ: ['ชื่อสินค้า','รุ่น','ราคา','รายละเอียด','alias','ชื่อที่มักถูกเรียก']
    """
    wb = _workbook()
    name_sheet = "ข้อมูลสินค้าและราคา"
    if name_sheet not in wb.sheetnames:
        return None
    ws = wb[name_sheet]
    items = _sheet_to_dicts(ws)
    if not items:
        return None

    first = items[0]
    name_col = _find_first_key(first, ["ชื่อสินค้า", "สินค้า", "รุ่น", "ชื่อรุ่น", "product_name"])
    price_col = _find_first_key(first, ["ราคา", "price"])
    desc_col = _find_first_key(first, ["รายละเอียด", "คำอธิบาย", "description"])
    alias_col = _find_first_key(first, ["alias", "ชื่อที่มักถูกเรียก", "ชื่อเรียก"])

    q = query.lower()
    best: Optional[Dict[str, Optional[str]]] = None

    # 1) ตรง alias ก่อน
    if alias_col:
        for row in items:
            aliases = safe_str(row.get(alias_col)).lower()
            if not aliases:
                continue
            if any(a and a in q for a in [s.strip() for s in aliases.split(",")]):
                best = row
                break

    # 2) ถ้าไม่เจอ ลองชื่อสินค้า
    if best is None and name_col:
        for row in items:
            name = safe_str(row.get(name_col)).lower()
            if name and name in q:
                best = row
                break

    if not best:
        return None

    name = safe_str(best.get(name_col)) if name_col else ""
    price = safe_str(best.get(price_col)) if price_col else ""
    desc = safe_str(best.get(desc_col)) if desc_col else ""

    parts = []
    if name:  parts.append(f"รุ่น: {name}")
    if price: parts.append(f"ราคา: {price}")
    if desc:  parts.append(f"รายละเอียด: {desc}")
    return "\n".join(parts) if parts else None

def get_payment_info() -> Optional[str]:
    wb = _workbook()
    if "Payment" not in wb.sheetnames:
        return None
    ws = wb["Payment"]
    items = _sheet_to_dicts(ws)
    if not items:
        return None

    # พยายามอ่านช่องทาง + รายละเอียด
    first = items[0]
    ch_col = _find_first_key(first, ["ช่องทาง", "method", "ชำระผ่าน"])
    dt_col = _find_first_key(first, ["รายละเอียด", "detail", "เงื่อนไข"])
    lines = []
    for r in items:
        ch = safe_str(r.get(ch_col)) if ch_col else ""
        dt = safe_str(r.get(dt_col)) if dt_col else ""
        if ch or dt:
            lines.append(f"• {ch}: {dt}" if ch else f"• {dt}")
    return "\n".join(lines) if lines else None

def get_after_sale_instruction() -> Optional[str]:
    wb = _workbook()
    name = "Intent Instruction – หลังการขาย"
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    items = _sheet_to_dicts(ws)
    if not items:
        return None
    # รวมข้อความทุกคอลัมน์ใน 5 แถวแรกเป็น guideline
    out: List[str] = []
    for row in items[:5]:
        for v in row.values():
            s = safe_str(v)
            if s:
                out.append(s)
    return "\n".join(out) if out else None

def get_pre_sale_instruction() -> Optional[str]:
    wb = _workbook()
    name = "Intent Instruction – ก่อนการขาย"
    if name not in wb.sheetnames:
        return None
    ws = wb[name]
    items = _sheet_to_dicts(ws)
    if not items:
        return None
    out: List[str] = []
    for row in items[:5]:
        for v in row.values():
            s = safe_str(v)
            if s:
                out.append(s)
    return "\n".join(out) if out else None
